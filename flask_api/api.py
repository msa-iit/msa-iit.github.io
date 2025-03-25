from flask import Flask, jsonify
from flask_cors import CORS
import dropbox
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, timedelta
import requests
import json


def getDropbox():
    """
    Retrieves Dropbox API key information from a JSON file and returns a
    Dropbox object initialized with the key information.
    
    :return: An instance of the Dropbox class initialized with the provided Dropbox API key, app key,
    app secret, and refresh token.
    """
    # Retrieve Dropbox API key necessary to login and access files on MSA Dropbox account
    with open("./data/API_keys.json", 'r') as json_file:
        data = json.load(json_file)
        APP_KEY = data["Dropbox_app_key"]
        APP_SECRET = data["Dropbox_app_secret"]
        REFRESH_TOKEN = data["Dropbox_refresh_token"]

    return dropbox.Dropbox(oauth2_refresh_token=REFRESH_TOKEN, app_key=APP_KEY, app_secret=APP_SECRET)


def format_time(time_str, input_format='%H:%M', output_format='%I:%M %p'):
    """
    The `format_time` function takes a time string in a specified input format, parses it, and returns
    the time string in a different output format. The Primary use is to convert from 24-hour to 12-hour format
    
    :param time_str: The `time_str` parameter is a string representing a time value in the format
    specified by the `input_format` parameter. It could be something like "14:30" if the input format is
    '%H:%M' (24-hour format)

    :param input_format: The `input_format` parameter specifies the format in which the `time_str` input
    is provided. By default, it is set to '%H:%M', which represents hours and minutes in 24-hour format,
    defaults to %H:%M (optional)

    :param output_format: The `output_format` parameter in the `format_time` function is used to specify
    the desired format in which the time should be returned after processing. In the provided function,
    the default `output_format` is set to `'%I:%M %p'`, which represents the time in 12, defaults to
    %I:%M %p (optional)

    :return: The function `format_time` is returning a formatted time string based on the input time
    string and formats provided.
    """
    # Parse the time string according to the input format
    try:
        time_str = time_str[0:5]
        time_obj = datetime.strptime(time_str, input_format)
    except ValueError as e:
        # Handle incorrect format or values
        print("API.format_time(): Invalid time format")
        print(e)
        return time_str
    
    # Convert and format the time string according to the output format
    return time_obj.strftime(output_format)

# Keep this. It is necessary to run Flask
app = Flask(__name__)
CORS(app)

@app.route('/LoadImages')
def LoadImages():
    """
    The function `LoadImages` retrieves web links to images hosted on Dropbox from a specified folder
    path.
    :return: The `LoadImages` function returns a list of web links to the images hosted on Dropbox after
    processing the files in the specified folder path.
    """

    with open("./data/MetaData.json", 'r') as file:
        IMAGES_FILE_PATH = json.load(file)["Images_Filepath"] # Path to folder that contains slideshow pictures in DROPBOX  

    dbx = getDropbox() # Get dropbox object to interact with files in dropbox account
    links = []  #Will contain the web links to the images hosted on dropbox
    entries = dbx.files_list_folder(IMAGES_FILE_PATH).entries
    for entry in entries:
        if isinstance(entry, dropbox.files.FileMetadata):
            try:
                # Try to get existing shared links for the file
                shared_links = dbx.sharing_list_shared_links(path=entry.path_lower).links
                # Check if there are any existing shared links
                if shared_links:
                    link = shared_links[0].url.replace("dl=0", "raw=1")  # Use the first existing link
                else:
                    # No existing links, create a new one
                    shared_link_metadata = dbx.sharing_create_shared_link_with_settings(entry.path_lower)
                    link = shared_link_metadata.url.replace("dl=0", "raw=1")
                links.append(link)
            except dropbox.exceptions.ApiError as api_err:
                print(f"LoadImages: Error obtaining link for {entry.path_lower}: {api_err}")
    return links

@app.route('/Iqamahs')
def Iqamahs():
    """
    The function `Iqamahs` checks if a file containing prayer times has been modified, downloads and
    processes the data if needed, and returns the prayer times in a JSON format.
    :return: The function `Iqamahs()` returns a JSON response containing the iqamah times for different
    prayers, converted to 12-hour format if they are in 24-hour format.
    """
    dbx = getDropbox()  # Get dropbox object to interact with files in dropbox account

    # Retrieve path to folder that contains Iqamahs excel file in DROPBOX
    with open("./data/MetaData.json", 'r') as file:
        IQAMAHS_FILE_PATH = json.load(file)["Iqamahs_Filepath"]

    # Get when what the last time the Iqamahs file was modified
    metadata = dbx.files_get_metadata(IQAMAHS_FILE_PATH)
    modified_date = metadata.server_modified

    # Get the saved modified date for the iqamahs file
    with open("./data/Dropbox_Last_Modified.json", 'r') as json_file:
        dropbox_json = json.load(json_file)
        saved_modified_date = dropbox_json['Iqamahs']

    # Compare the dates. If they do not equal then the file was modified
    if saved_modified_date != str(modified_date):
        # Store the new modified date of the iqamahs file. This will be the new date checked in the future to see if the file was changed again
        with open("./data/Dropbox_Last_Modified.json", 'w') as json_file:
            dropbox_json['Iqamahs'] = str(modified_date)
            json.dump(dropbox_json, json_file)

        # Download the file
        _, res = dbx.files_download(IQAMAHS_FILE_PATH)
        
        # Load the workbook
        workbook = load_workbook(filename=BytesIO(res.content))
        
        # Assuming your data is in the first sheet
        first_sheet_name = workbook.sheetnames[0]
        sheet = workbook[first_sheet_name]
        
        # Construct json object for iqamahs data
        iqamahs = {
            "Fajr": sheet['B2'].value,
            "Dhuhr": sheet['B3'].value,
            "Asr": sheet['B4'].value,
            "Maghrib": sheet['B5'].value,
            "Isha": sheet['B6'].value,
            "Jummah Khutbah": sheet['B7'].value,
            "Jummah Iqamah": sheet['B8'].value
        }
        
        # Save the new iqamahs data
        with open("./data/Iqamahs_Save.json", 'w') as json_file:
            json.dump(iqamahs, json_file)

    else:
        #Dates match meaning the file was not changed. Load up old data
        with open("./data/Iqamahs_Save.json", 'r') as json_file:
            iqamahs = json.load(json_file)

    for key in iqamahs.keys():
        # Any prayers that do not have an Iqamah assigned, convert to blank string ""
        if not iqamahs[key]:
            iqamahs[key] = ""
        else:
            # Any prayers that do have an Iqamah time, convert from 24-hour format to 12-hour format and save
            iqamahs[key] = format_time(str(iqamahs[key]))

    return jsonify(iqamahs)

@app.route('/SlideshowDelay')
def slideshowDelay():
    """
    The function `slideshowDelay` checks for updates in a Dropbox file containing TV settings, retrieves
    the slideshow delay value, and updates it in a JSON file.
    :return: The function `slideshowDelay()` returns the slideshow delay value as a string.
    """
    dbx = getDropbox() # Get dropbox object to interact with files in dropbox account

    # Retrieve path to folder that contains TV Settings excel file in DROPBOX
    with open("./data/MetaData.json", 'r') as file:
        TV_SETTINGS_FILE_PATH = json.load(file)["TV_settings_Filepath"]

    # Get metadata for the file
    metadata = dbx.files_get_metadata(TV_SETTINGS_FILE_PATH)
    modified_date = metadata.server_modified

    # Get saved modified date
    with open("./data/Dropbox_Last_Modified.json", 'r') as json_file:
        dropbox_json = json.load(json_file)
        saved_modified_date = dropbox_json['TV Settings']

    # Compare the dates
    if saved_modified_date != str(modified_date):
        # Update the saved modified date
        with open("./data/Dropbox_Last_Modified.json", 'w') as json_file:
            dropbox_json['TV Settings'] = str(modified_date)
            json.dump(dropbox_json, json_file)

        # Download the file
        _, res = dbx.files_download(TV_SETTINGS_FILE_PATH)
        
        # Load the workbook
        workbook = load_workbook(filename=BytesIO(res.content))
        
        # Assuming your data is in the first sheet, adjust as necessary
        first_sheet_name = workbook.sheetnames[0]
        sheet = workbook[first_sheet_name]
        
        # Construct json object for iqamahs data
        for i in range(2,50):
            if sheet[f'A{i}'].value == "Slideshow Delay":
                try:
                    if sheet[f'B{i}'].value:
                        delay = int(sheet[f'B{i}'].value)
                    else:
                        print("Value for 'Slideshow Delay' found in excel file in Dropbox does not appear to have a value assigned. Defaulting to 15s delay")
                        delay = 15
                except ValueError:
                    print("Value for 'Slideshow Delay' found in excel file in Dropbox does not appear to be an integer. Defaulting to 15s delay")
                    delay = 15
                break
        
        #Retrieve json of TV Settings as a dictionary, update the slideshow delay, and save settings back into a file
        with open("./data/TV_Settings_Save.json", 'r') as json_file:
            settings = json.load(json_file)
        
        settings["Slideshow Delay"] = delay

        with open("./data/TV_Settings_Save.json", 'w') as json_file:
            json.dump(settings, json_file)

    else:
        # Dates match which menas file was not modified. Use the saved slideshow delay
        with open("./data/TV_Settings_Save.json", 'r') as json_file:
            delay = json.load(json_file)["Slideshow Delay"]

    return str(delay)

@app.route('/prayerAPI')
def prayerAPI():
    """
    Retrieves prayer time data from an API, saves it locally, and returns the data
    either from the API or the saved file based on the last update date.
    :return: The code is returning the prayer times API response either by making a new API call to
    fetch the data for the current month and year if the data is not already saved, or by returning the
    saved API response if the data for the current month and year is already available.
    """

    # Grab the current month and year
    date = datetime.now()
    with open('./data/MetaData.json') as f: 
        file_content = json.load(f) 
    month = int(file_content["last_aladhanAPI_call"][:2]) 
    year = int(file_content["last_aladhanAPI_call"][3:7])

    # If the current month or year exceeds the previous month or year, then retrieve the new prayer times from the aladhan api.
    if date.year > year or (date.month > month): 
        method = 2
        school = 1 
        api_response = requests.get(f"https://api.aladhan.com/v1/calendarByCity/{date.year}/{date.month}?city=Chicago&country=United%20States&method={method}&school={school}").json() 
        with open('./data/aladhanAPIsave.json', 'w') as f: 
            json.dump(api_response, f) # Update Settings File to remember that we grabbed new data for this month 
        file_content["last_aladhanAPI_call"] = f"{date.month:02d}/{date.year}" 
        with open('./data/MetaData.json', 'w') as f: 
            json.dump(file_content, f) 
            return jsonify(api_response)
    else: 
        # Retrieve the saved prayer times from the file
        with open('./data/aladhanAPIsave.json', encoding='utf-8') as f:
            api_save = json.load(f)
        return jsonify(api_save)

@app.route('/prayerTimesToday')
def prayerTimesToday():
    """
    Returns the prayer times for today
    """
    
    jsonAPIdata = dict(prayerAPI().json)    # Retrieve prayer times for the whole month from the prayer times API save file
    today = datetime.now()
    idx = today.day - 1

    today_data = jsonAPIdata['data'][idx]['timings']    # Get only the prayer timings for today

    # Convert 24-hour format to 12-hour format
    for key in today_data.keys():
        today_data[key] = format_time(str(today_data[key]))

    return jsonify(today_data)

@app.route('/todayHijri')
def todayHijri():
    """
    Retrieve the Hijri Date for today. 
    Most of this method is just trying to parse through the prayer API save file for the weekday, day, month, and year
    """
    jsonAPIdata = dict(prayerAPI().json)    # Prayer times for the month from the prayerAPI() function
    MONTHS = ["Muharram", "Safar", "Rabi al-Awwal", "Rabi al-Thani", "Jumada al-Awwal", "Jumada al-Thani", "Rajab", "Shaban", "Ramadan", "Shawwal", "Dhu al-Qadah", "Dhu al-Hijjah"];
    today = datetime.now()
    day_idx = today.day - 1 # The index within the json data that contains today's information

    weekday = jsonAPIdata['data'][day_idx]['date']['hijri']['weekday']['en']
    day = jsonAPIdata['data'][day_idx]['date']['hijri']['day']
    month_idx = int(jsonAPIdata['data'][day_idx]['date']['hijri']['month']['number']) - 1
    year = jsonAPIdata['data'][day_idx]['date']['hijri']['year']
    month = MONTHS[month_idx]

    return str(weekday + ", " + month + " " + day + ", " + year)

@app.route('/NextSalah')
def NextSalah():
    """
    Return which Salah is next and what time it will be at. Used to calculate the end for the countdown timer on the website TV display
    """

    # Get Iqamah times and convert them from 12-hour format back to 24-hour format to make it easier to calculate which iqamah/prayer is next
    Iqamah_Times = dict(Iqamahs().json)
    for key in Iqamah_Times.keys():
        if Iqamah_Times[key] != "":
            Iqamah_Times[key] = datetime.strptime(Iqamah_Times[key], "%I:%M %p").strftime("%H:%M")

    # Get Prayer times and convert them from 12-hour format back to 24-hour format to make it easier to calculate which iqamah/prayer is next
    Today_Times = dict(prayerTimesToday().json)
    for key in Today_Times.keys():
        if Today_Times[key] != "":
            Today_Times[key] = datetime.strptime(Today_Times[key], "%I:%M %p").strftime("%H:%M")

    # Convert all the important timings into python datetime for easier calculation
    currentTime = datetime.now()
    StartOfDay = datetime(currentTime.year, currentTime.month, currentTime.day, 0, 0, 0, 0)

    FajrHour = int(Today_Times['Fajr'][:2])
    FajrMinute = int(Today_Times['Fajr'][3:5])
    FajrTime = datetime(currentTime.year, currentTime.month, currentTime.day, FajrHour, FajrMinute, 0, 0)

    FajrIqamahHour = int(Iqamah_Times['Fajr'][:2]) if Iqamah_Times['Fajr'] else None
    FajrIqamahMinute = int(Iqamah_Times['Fajr'][3:5]) if Iqamah_Times['Fajr'] else None
    FajrIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, FajrHour, FajrMinute, 0, 0) if FajrIqamahHour and FajrIqamahMinute else None

    SunriseHour = int(Today_Times['Sunrise'][:2])
    SunriseMinute = int(Today_Times['Sunrise'][3:5])
    SunriseTime = datetime(currentTime.year, currentTime.month, currentTime.day, SunriseHour, SunriseMinute, 0, 0)

    DhuhrHour = int(Today_Times['Dhuhr'][:2])
    DhuhrMinute = int(Today_Times['Dhuhr'][3:5])
    DhuhrTime = datetime(currentTime.year, currentTime.month, currentTime.day, DhuhrHour, DhuhrMinute, 0, 0)

    DhuhrIqamahHour = int(Iqamah_Times['Dhuhr'][:2]) if len(Iqamah_Times['Dhuhr']) else None
    DhuhrIqamahMinute = int(Iqamah_Times['Dhuhr'][3:5]) if len(Iqamah_Times['Dhuhr']) else None
    DhuhrIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, DhuhrIqamahHour, DhuhrIqamahMinute, 0, 0) if DhuhrIqamahHour and DhuhrIqamahMinute else None
    JummahKhutbahHour = int(Iqamah_Times['Jummah Khutbah'][:2])
    JummahKhutbahMinute = int(Iqamah_Times['Jummah Khutbah'][3:5])
    JummahKhutbah = datetime(currentTime.year, currentTime.month, currentTime.day, JummahKhutbahHour, JummahKhutbahMinute, 0, 0)

    JummahIqamahHour = int(Iqamah_Times['Jummah Iqamah'][:2])
    JummahIqamahMinute = int(Iqamah_Times['Jummah Iqamah'][3:5])
    JummahIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, JummahIqamahHour, JummahIqamahMinute, 0, 0)

    AsrHour = int(Today_Times['Asr'][:2])
    AsrMinute = int(Today_Times['Asr'][3:5])
    AsrTime = datetime(currentTime.year, currentTime.month, currentTime.day, AsrHour, AsrMinute, 0, 0)

    AsrIqamahHour = int(Iqamah_Times['Asr'][:2]) if len(Iqamah_Times['Asr']) else None
    AsrIqamahMinute = int(Iqamah_Times['Asr'][3:5]) if len(Iqamah_Times['Asr']) else None
    AsrIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, AsrIqamahHour, AsrIqamahMinute, 0, 0) if AsrIqamahHour and AsrIqamahMinute else None

    MaghribHour = int(Today_Times['Maghrib'][:2])
    MaghribMinute = int(Today_Times['Maghrib'][3:5])
    MaghribTime = datetime(currentTime.year, currentTime.month, currentTime.day, MaghribHour, MaghribMinute, 0, 0)

    MaghribIqamahHour = int(Iqamah_Times['Maghrib'][:2]) if len(Iqamah_Times['Maghrib']) else None
    MaghribIqamahMinute = int(Iqamah_Times['Maghrib'][3:5]) if len(Iqamah_Times['Maghrib']) else None
    MaghribIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, MaghribIqamahHour, MaghribIqamahMinute, 0, 0) if MaghribIqamahHour and MaghribIqamahMinute else None

    IshaHour = int(Today_Times['Isha'][:2])
    IshaMinute = int(Today_Times['Isha'][3:5])
    IshaTime = datetime(currentTime.year, currentTime.month, currentTime.day, IshaHour, IshaMinute, 0, 0)

    IshaIqamahHour = int(Iqamah_Times['Isha'][:2]) if len(Iqamah_Times['Isha']) else None
    IshaIqamahMinute = int(Iqamah_Times['Isha'][3:5]) if len(Iqamah_Times['Isha']) else None
    IshaIqamah = datetime(currentTime.year, currentTime.month, currentTime.day, IshaIqamahHour, IshaIqamahMinute, 0, 0) if IshaIqamahHour and IshaIqamahMinute else None

    # Chain of if-else statements checking which timing is next to display
    if StartOfDay <= currentTime <= FajrTime:
        result = {
            'prayer': 'Fajr',
            'type': 'time',
            'time': FajrTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif FajrIqamah and FajrTime <= currentTime < FajrIqamah:
        result = {
            'prayer': 'Fajr',
            'type': 'iqamah',
            'time': FajrIqamah.strftime("%a, %d %b %Y %H:%M")
        }
    elif FajrTime <= currentTime < SunriseTime:
        result = {
            'prayer': 'Sunrise',
            'type': 'time',
            'time': SunriseTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif currentTime.weekday() == 4 and SunriseTime <= currentTime < JummahKhutbah:
        result = {
            'prayer': 'Jummah',
            'type': 'khutbah',
            'time': JummahKhutbah.strftime("%a, %d %b %Y %H:%M")
        }
    elif currentTime.weekday() == 4 and JummahKhutbah <= currentTime < JummahIqamah:
        result = {
            'prayer': 'Jummah',
            'type': 'iqamah',
            'time': JummahIqamah.strftime("%a, %d %b %Y %H:%M")
        }    
    elif currentTime.weekday() == 4 and JummahIqamah <= currentTime < AsrTime:
        result = {
            'prayer': 'Asr',
            'type': 'time',
            'time': AsrTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif SunriseTime <= currentTime < DhuhrTime:
        result = {
            'prayer': 'Dhuhr',
            'type': 'time',
            'time': DhuhrTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif DhuhrIqamah and DhuhrTime <= currentTime < DhuhrIqamah:
        result = {
            'prayer': 'Dhuhr',
            'type': 'iqamah',
            'time': DhuhrIqamah.strftime("%a, %d %b %Y %H:%M")
        }
    elif DhuhrTime <= currentTime < AsrTime: 
        result = { 
            'prayer': 'Asr',
            'type': 'time', 
            'time': AsrTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif AsrIqamah and AsrTime <= currentTime < AsrIqamah:
        result = {
            'prayer': 'Asr',
            'type': 'iqamah',
            'time': AsrIqamah.strftime("%a, %d %b %Y %H:%M")
        }
    elif AsrTime <= currentTime < MaghribTime: 
        result = { 
            'prayer': 'Maghrib',
            'type': 'time', 
            'time': MaghribTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif MaghribIqamah and MaghribTime <= currentTime < MaghribIqamah:
        result = {
            'prayer': 'Maghrib',
            'type': 'iqamah',
            'time': MaghribIqamah.strftime("%a, %d %b %Y %H:%M")
        }
    elif MaghribTime <= currentTime < IshaTime: 
        result = { 
            'prayer': 'Isha',
            'type': 'time', 
            'time': IshaTime.strftime("%a, %d %b %Y %H:%M")
        }
    elif IshaIqamah and IshaTime <= currentTime < IshaIqamah:
        result = {
            'prayer': 'Isha',
            'type': 'iqamah',
            'time': IshaIqamah.strftime("%a, %d %b %Y %H:%M")
        }
    else:
        result = { 
            'prayer': 'Fajr',
            'type': 'time', 
            'time': (FajrTime + timedelta(days=1)).strftime("%a, %d %b %Y %H:%M")
        }

    return jsonify(result)

@app.route('/todayGreg')
def todayGreg():
    """
    Returns the Gregorian date for today (The regular old date everyone is used to)
    """
    return datetime.now().strftime('%A, %B %d, %Y')

#Initialize Flask and run on port 7000
if __name__ == '__main__':
    app.run(port=7000)